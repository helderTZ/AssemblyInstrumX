import sys
from openpyxl import *

def showList(l):
    for i in range(len(l)):
        l[i].show()

class InstructionSet():
    def __init__(self, name):
        self.name = name
        self.nbInstructions = 0
        self.instructions = []

    def addInstruction(self, inst):
        self.instructions.append(inst)
        self.nbInstructions += 1
        inst.set = self
    
    def show(self, f=sys.stdout):
        print(self.name + ' ' + str(self.nbInstructions), file=f)
    
    def deep_show(self, f=sys.stdout):
        print(self.name + ' ' + str(self.nbInstructions), file=f)
        for i in self.instructions:
            print('\t' + i.show(), file=f)

class Instruction():
    def __init__(self, name):
        self.opcode = name
        self.sufix = ''
        self.operands = []
        self.nbOperands = 0
        self.set = None
        self.labels = []
        self.nbLabels = 0
        self.dataMode = ''
        self.dataType = ''
        self.dataTraffic = 0
    
    def addOperand(self, name):
        self.operands.append(name)
        self.nbOperands += 1
    
    def addLabel(self, label):
        self.labels.append(label)
        self.nbLabels += 1
    
    def show(self):
        desc = str(self.opcode)
        desc += '\t' + str(self.set.name)
        for op in self.operands:
            desc += '\t' + str(op.show())
        desc += '\t'
        for label in self.labels:
            desc += str(label) + '/'
        desc = desc[:-1]
        desc += '\t' + str(self.dataMode)
        desc += '\t' + str(self.dataType)
        desc += '\t' + str(self.dataTraffic)
        return desc
        
        
def isReg(reg):
    if ('(' in reg):
        return False
    else:
        return True

class Operand():
    def __init__(self, name):
        self.name = name
        self.reg = isReg(name)
    
    def show(self):
        return str(self.name)




def fill_instruction_database():
    # DEBUG
    f = open('cells.txt', 'w')
    
    instructionSets = []
    instructions = []
    wb = load_workbook('Intel-Instructions.xlsx')
    sheet_names = wb.get_sheet_names()
    setIdx = 0
    
    # iterate through worksheets
    for ws_name in sheet_names:
        ws = wb.get_sheet_by_name(ws_name)
        newInstructionSet = InstructionSet(ws_name)
        instructionSets.append( newInstructionSet )
        
        # iterate through rows over column B
        col = "A"
        for row in range(2, 300):
            cell_name = '{}{}'.format(col, row)
            cell_content = ws[cell_name].value
            cell_labels_name = '{}{}'.format('B', row)
            cell_labels_content = ws[cell_labels_name].value
            cell_mode_name = '{}{}'.format('C', row)
            cell_mode_content = ws[cell_mode_name].value
            cell_type_name = '{}{}'.format('D', row)
            cell_type_content = ws[cell_type_name].value
            cell_bits_name = '{}{}'.format('E', row)
            cell_bits_content = ws[cell_bits_name].value
    
            # empty line
            if (cell_content == '' or cell_content == None):
                continue
            if (cell_labels_content == None):
                cell_labels_content = ''
            if (cell_mode_content == None):
                cell_mode_content = ''
            if (cell_type_content == None):
                cell_type_content = ''
            
            # DEBUG
            print(cell_content, file=f)
            
            cell_content = cell_content.split(' ')
            inst_name = cell_content[0]
    
            # get operands if available
            operands = []
            if (len(cell_content) > 1):
                cell_content = cell_content[1:]
                for i in range(len(cell_content)):
                    operandName = cell_content[i].replace(',', '')
                    operand = Operand(operandName)
                    operands.append(operand)
                    
            # separate instructions in same cell separated by '/'
            if ('/' in inst_name):
                newInstructionNames = inst_name.split('/')
            else:
                newInstructionNames = inst_name.split(' ')
            
            for name in newInstructionNames:
                newInst = Instruction(name)
                instructionSets[setIdx].addInstruction(newInst)
                instructions.append( newInst )
                for op in operands:
                    newInst.addOperand(op)
                labels = cell_labels_content.split('/')
                for l in labels:
                    newInst.addLabel(l)
                newInst.dataMode = cell_mode_content
                newInst.dataType = cell_type_content
                newInst.dataTraffic = cell_bits_content
                
            
        setIdx += 1
    
    # DEBUG
    f.close()
    
    return [instructionSets, instructions]

def get_instruction_names(instructions):
    names = []
    for i in instructions:
        names.append(i.opcode)
    return names

def main():
    [instructionSets, instructions] = fill_instruction_database()
    with open('output.txt', 'w') as f:
        for i in instructionSets:
            i.deep_show(f)

if __name__ == "__main__":
    main()
